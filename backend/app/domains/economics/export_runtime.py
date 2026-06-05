from __future__ import annotations

import csv
import io
import json
from collections import Counter
from dataclasses import dataclass
from datetime import date, datetime, timezone
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.core.config import get_settings
from app.core.logging import get_logger
from app.domains.cloud_accounts.service import CloudAccountService
from app.domains.cloud_ledger.service import CloudLedgerService
from app.domains.cloud_ledger.schemas import SubscriptionCostSummary
from app.domains.decision_engine.service import DecisionEngineService
from app.domains.economics.models import ReportExportFormat, ReportExportJob

log = get_logger(__name__)


@dataclass
class ReportExportArtifact:
    file_name: str
    content_type: str
    content: bytes


def ensure_report_exports_dir() -> Path:
    export_dir = Path(get_settings().report_exports_dir)
    export_dir.mkdir(parents=True, exist_ok=True)
    return export_dir


def persist_report_export_file(job_id: str, file_format: ReportExportFormat, content: bytes) -> Path:
    export_dir = ensure_report_exports_dir()
    export_path = export_dir / f"{job_id}.{file_format.value}"
    export_path.write_bytes(content)
    return export_path.resolve()


def _escape_xml(text: str) -> str:
    return (
        text.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
        .replace("'", "&apos;")
    )


def _cell_ref(column_index: int, row_index: int) -> str:
    value = ""
    current = column_index
    while current > 0:
        current, remainder = divmod(current - 1, 26)
        value = chr(65 + remainder) + value
    return f"{value}{row_index}"


def _worksheet_xml(
    rows: list[list[object]],
    *,
    header_row: int = 1,
    col_formats: dict[int, int] | None = None,
) -> str:
    col_formats = col_formats or {}
    row_xml: list[str] = []
    for row_index, row in enumerate(rows, start=1):
        cells: list[str] = []
        for column_index, value in enumerate(row, start=1):
            ref = _cell_ref(column_index, row_index)
            style_id = 0
            if row_index == header_row:
                style_id = 1  # bold
            elif column_index in col_formats:
                style_id = col_formats[column_index]
            if value is None:
                cells.append(f'<c r="{ref}" s="{style_id}" t="inlineStr"><is><t></t></is></c>')
            elif isinstance(value, (int, float)) and not isinstance(value, bool):
                cells.append(f'<c r="{ref}" s="{style_id}"><v>{value}</v></c>')
            else:
                text = _escape_xml(str(value))
                cells.append(f'<c r="{ref}" s="{style_id}" t="inlineStr"><is><t>{text}</t></is></c>')
        row_xml.append(f'<row r="{row_index}">{"".join(cells)}</row>')
    freeze = ""
    if header_row >= 1:
        freeze_ref = _cell_ref(1, header_row + 1)
        freeze = (
            '<sheetViews><sheetView tabSelected="1" workbookViewId="0">'
            f'<pane ySplit="{header_row}" topLeftCell="{freeze_ref}" activePane="bottomLeft" state="frozen"/>'
            '</sheetView></sheetViews>'
        )
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        f'{freeze}'
        f'<sheetData>{"".join(row_xml)}</sheetData>'
        '</worksheet>'
    )


@dataclass
class SheetDef:
    name: str
    rows: list[list[object]]
    header_row: int = 1
    col_formats: dict[int, int] | None = None


def build_xlsx_workbook(sheets: list[tuple[str, list[list[object]]]] | list[SheetDef]) -> bytes:
    sheet_defs: list[SheetDef] = []
    for item in sheets:
        if isinstance(item, SheetDef):
            sheet_defs.append(item)
        else:
            name, rows = item
            sheet_defs.append(SheetDef(name=name, rows=rows))

    buffer = io.BytesIO()
    with ZipFile(buffer, "w", compression=ZIP_DEFLATED) as zf:
        zf.writestr(
            "[Content_Types].xml",
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
            '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
            '<Default Extension="xml" ContentType="application/xml"/>'
            '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
            '<Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>'
            + "".join(
                f'<Override PartName="/xl/worksheets/sheet{idx}.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
                for idx in range(1, len(sheet_defs) + 1)
            )
            + '</Types>',
        )
        zf.writestr(
            "_rels/.rels",
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>'
            '</Relationships>',
        )
        zf.writestr(
            "xl/workbook.xml",
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
            'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
            '<sheets>'
            + "".join(
                f'<sheet name="{_escape_xml(sd.name[:31])}" sheetId="{idx}" r:id="rId{idx}"/>'
                for idx, sd in enumerate(sheet_defs, start=1)
            )
            + '</sheets></workbook>',
        )
        zf.writestr(
            "xl/_rels/workbook.xml.rels",
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            + "".join(
                f'<Relationship Id="rId{idx}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet{idx}.xml"/>'
                for idx in range(1, len(sheet_defs) + 1)
            )
            + f'<Relationship Id="rId{len(sheet_defs) + 1}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>'
            + '</Relationships>',
        )
        # Styles: 0=normal, 1=bold header, 2=currency ($#,##0.00), 3=percentage (0.0%)
        zf.writestr(
            "xl/styles.xml",
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
            '<numFmts count="2">'
            '<numFmt numFmtId="164" formatCode="&quot;$&quot;#,##0.00"/>'
            '<numFmt numFmtId="165" formatCode="0.0%"/>'
            '</numFmts>'
            '<fonts count="2">'
            '<font><sz val="11"/><name val="Calibri"/></font>'
            '<font><b/><sz val="11"/><name val="Calibri"/></font>'
            '</fonts>'
            '<fills count="2"><fill><patternFill patternType="none"/></fill><fill><patternFill patternType="gray125"/></fill></fills>'
            '<borders count="1"><border/></borders>'
            '<cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>'
            '<cellXfs count="4">'
            '<xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/>'
            '<xf numFmtId="0" fontId="1" fillId="0" borderId="0" xfId="0" applyFont="1"/>'
            '<xf numFmtId="164" fontId="0" fillId="0" borderId="0" xfId="0" applyNumberFormat="1"/>'
            '<xf numFmtId="165" fontId="0" fillId="0" borderId="0" xfId="0" applyNumberFormat="1"/>'
            '</cellXfs>'
            '</styleSheet>',
        )
        for idx, sd in enumerate(sheet_defs, start=1):
            zf.writestr(
                f"xl/worksheets/sheet{idx}.xml",
                _worksheet_xml(sd.rows, header_row=sd.header_row, col_formats=sd.col_formats),
            )
    return buffer.getvalue()


_CATEGORY_LABELS = {
    "rightsizing": "Redimensionamento",
    "aks_nodepool_rightsizing": "AKS Node Pool",
    "aks_autoscaler_recommendation": "AKS Autoscaler",
    "idle_resources": "Recursos Ociosos",
    "reserved_instances": "Instâncias Reservadas",
    "storage_optimization": "Otimização de Storage",
    "network_optimization": "Otimização de Rede",
    "license_optimization": "Otimização de Licenças",
    "architecture_change": "Mudança de Arquitetura",
}

_STATUS_LABELS = {
    "open": "Aberta",
    "in_progress": "Em Progresso",
    "resolved": "Resolvida",
    "dismissed": "Descartada",
    "validated": "Validada",
}


def _format_category(raw: str) -> str:
    return _CATEGORY_LABELS.get(raw, raw.replace("_", " ").title())


def _format_status(raw: str) -> str:
    return _STATUS_LABELS.get(raw, raw.replace("_", " ").title())


def _format_date_br(value) -> str:
    """Format date as DD/MM/YYYY for Brazilian Excel."""
    if isinstance(value, (date, datetime)):
        return value.strftime("%d/%m/%Y")
    return str(value) if value else ""


def _format_datetime_br(value) -> str:
    """Format datetime as DD/MM/YYYY HH:MM for Brazilian Excel."""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y %H:%M")
    elif isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    return str(value) if value else ""


async def build_report_export_artifact(db, job: ReportExportJob) -> ReportExportArtifact:
    account_service = CloudAccountService(db)
    accounts, _ = await account_service.list_accounts(job.org_id)
    active_accounts = sum(1 for account in accounts if account.status.value == "active")

    ledger = CloudLedgerService(db)
    dashboard = await ledger.get_dashboard_metrics(job.org_id, active_accounts)
    top_services, _ = ledger.get_top_services(job.org_id, days=job.window_days, limit=15)
    top_teams, _ = ledger.get_top_teams(job.org_id, days=job.window_days, limit=15)
    trend = ledger.get_cost_trend(job.org_id, days=job.window_days)

    # Get subscription breakdown (NEW)
    subscription_summary: SubscriptionCostSummary = await ledger.get_subscription_cost_breakdown(
        job.org_id, days=job.window_days
    )

    decision_engine = DecisionEngineService(db)
    opportunities, _ = await decision_engine.list_opportunities(job.org_id, limit=50)

    # Defensive: ensure all iterables are never None
    opportunities = opportunities or []

    total_monthly_savings = sum(op.estimated_monthly_savings_usd for op in opportunities)
    total_annual_savings = sum(op.estimated_annual_savings_usd for op in opportunities)

    generated_at = datetime.now(timezone.utc)
    filters_json = json.dumps(job.filters or {}, sort_keys=True)
    base_name = f"causium-relatorio-finops-{generated_at.strftime('%Y%m%d-%H%M%S')}"

    # Get org name from first account or job
    org_name = "N/A"
    if accounts:
        org_name = str(accounts[0].org_id)[:8]

    # Fetch governance + sustainability data (used by both CSV and XLSX)
    from app.domains.gov.service import GovService
    from app.domains.green.service import GreenService

    gov = GovService()
    gov_summary = gov.get_summary(job.org_id, days=job.window_days)
    gov_unowned = gov.get_unowned_costs(job.org_id, days=job.window_days, limit=30) or []
    gov_compliance = gov.get_label_compliance(job.org_id, days=job.window_days) or []

    green = GreenService()
    green_summary = green.get_summary(job.org_id, months=6)
    green_monthly = green.get_emissions_monthly(job.org_id, months=6) or []

    if job.file_format == ReportExportFormat.CSV:
        content = _build_csv_zip(
            dashboard=dashboard,
            top_services=top_services,
            top_teams=top_teams,
            trend=trend,
            opportunities=opportunities,
            total_monthly_savings=total_monthly_savings,
            total_annual_savings=total_annual_savings,
            subscription_summary=subscription_summary,
            gov_summary=gov_summary,
            gov_unowned=gov_unowned,
            gov_compliance=gov_compliance,
            green_summary=green_summary,
            green_monthly=green_monthly,
            generated_at=generated_at,
            window_days=job.window_days,
            org_name=org_name,
        )
        return ReportExportArtifact(
            file_name=f"{base_name}.zip",
            content_type="application/zip",
            content=content,
        )

    # --- XLSX via openpyxl ---
    workbook_bytes = _build_professional_xlsx(
        dashboard=dashboard,
        top_services=top_services,
        top_teams=top_teams,
        trend=trend,
        opportunities=opportunities,
        total_monthly_savings=total_monthly_savings,
        total_annual_savings=total_annual_savings,
        subscription_summary=subscription_summary,
        gov_summary=gov_summary,
        gov_unowned=gov_unowned,
        gov_compliance=gov_compliance,
        green_summary=green_summary,
        green_monthly=green_monthly,
        generated_at=generated_at,
        window_days=job.window_days,
        filters_json=filters_json,
        org_name=org_name,
    )
    return ReportExportArtifact(
        file_name=f"{base_name}.xlsx",
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        content=workbook_bytes,
    )


# ---------------------------------------------------------------------------
# CSV ZIP builder (improved for Brazilian Excel)
# ---------------------------------------------------------------------------

def _csv_bytes_br(headers: list[str], rows: list[list]) -> bytes:
    """Generate CSV with BOM UTF-8 and semicolon delimiter for Brazilian Excel."""
    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=";", quotechar='"', lineterminator="\r\n")
    writer.writerow(headers)
    for row in rows:
        # Format dates and clean values
        formatted_row = []
        for val in row:
            if isinstance(val, (date, datetime)):
                formatted_row.append(_format_date_br(val))
            elif val is None:
                formatted_row.append("")
            else:
                formatted_row.append(str(val))
        writer.writerow(formatted_row)
    # BOM UTF-8 for Excel compatibility
    return "﻿".encode("utf-8") + buf.getvalue().encode("utf-8")


def _build_csv_zip(
    *,
    dashboard,
    top_services,
    top_teams,
    trend,
    opportunities,
    total_monthly_savings: float,
    total_annual_savings: float,
    subscription_summary,
    gov_summary,
    gov_unowned,
    gov_compliance,
    green_summary,
    green_monthly,
    generated_at: datetime,
    window_days: int,
    org_name: str,
) -> bytes:
    # Defensive: ensure all optional iterables are never None
    top_services = top_services or []
    top_teams = top_teams or []
    trend = trend or []
    opportunities = opportunities or []
    gov_unowned = gov_unowned or []
    gov_compliance = gov_compliance or []
    green_monthly = green_monthly or []

    # 1. Resumo Executivo
    summary_rows = [
        ["Gasto Mensal Atual", round(dashboard.current_month_cost, 2)],
        ["Gasto Mês Anterior", round(dashboard.previous_month_cost, 2)],
        ["Variação MoM (%)", round(dashboard.mom_change_pct, 1)],
        ["Economia Mensal Potencial", round(total_monthly_savings, 2)],
        ["Economia Anual Potencial", round(total_annual_savings, 2)],
        ["Contas Cloud Ativas", dashboard.active_accounts],
        ["Oportunidades Abertas", len(opportunities)],
        ["Eventos (Últimos 7 dias)", dashboard.event_count_7d],
        ["Período do Relatório (dias)", window_days],
        ["Data de Geração", _format_datetime_br(generated_at)],
        ["Organização", org_name],
    ]
    summary_csv = _csv_bytes_br(["Indicador", "Valor"], summary_rows)

    # 2. Custos por Serviço (inline para resources.csv)
    res_rows = []
    if top_services:
        total_svc_cost = sum(s.cost_usd for s in top_services)
        for svc in top_services:
            pct = (svc.cost_usd / total_svc_cost * 100) if total_svc_cost > 0 else 0
            res_rows.append([
                svc.service,
                "Recurso principal",
                round(svc.cost_usd, 2),
                round(pct, 2),
            ])

    # 3. Custos por Equipe (inline)
    team_rows = []
    for team in top_teams:
        team_rows.append([
            team.service,
            round(team.cost_usd, 2),
            round(team.percentage, 1),
        ])

    # 4. Tendência Diária (inline)
    trend_rows = []
    for row in trend:
        trend_rows.append([
            _format_date_br(row.date),
            round(row.cost_usd, 2),
            row.provider or "Todos",
        ])

    # 5. Oportunidades (inline para recommendations.csv)
    rec_rows = []
    if opportunities:
        for i, op in enumerate(sorted(opportunities, key=lambda x: x.estimated_annual_savings_usd, reverse=True)[:20], 1):
            rec_rows.append([
                i,
                op.title,
                _format_category(op.category.value),
                round(op.estimated_annual_savings_usd, 2),
                op.risk_level.value.capitalize(),
                f"Implementar {_format_category(op.category.value).lower()}",
            ])

    # 6. Governança (defensive: handle None)
    gov_rows = []
    if gov_unowned:
        for item in gov_unowned:
            gov_rows.append([
                item.service,
                item.resource_id,
                item.region,
                item.environment,
                round(item.cost_usd, 2),
                item.days_active,
            ])
    gov_compliance_rows = []
    if gov_compliance:
        for item in gov_compliance:
            gov_compliance_rows.append([
                item.team,
                round(item.total_cost_usd, 2),
                round(item.untagged_cost_usd, 2),
                round(item.compliance_pct, 1),
            ])
    gov_csv = _csv_bytes_br(
        ["Serviço", "Recurso", "Região", "Ambiente", "Custo sem Dono", "Dias Ativo"],
        gov_rows,
    )
    gov_compliance_csv = _csv_bytes_br(
        ["Equipe", "Custo Total", "Custo sem Tag", "Compliance (%)"],
        gov_compliance_rows,
    )
    # Merge governance into one file with a section separator
    gov_combined = gov_csv + "\r\n".encode("utf-8") + b"SECAO:COMPLIANCE POR EQUIPE\r\n" + gov_compliance_csv

    # 7. Sustentabilidade (defensive: handle None)
    green_rows = []
    if green_monthly:
        for m in green_monthly:
            green_rows.append([
                m.month,
                round(m.kg_co2e, 1),
                round(m.cost_usd, 2),
                round(m.delta_pct, 1) if m.delta_pct else "",
            ])
    green_csv = _csv_bytes_br(["Mês", "Emissões (kgCO2e)", "Custo", "Variação (%)"], green_rows)

    # 8. Subscriptions (CRITICAL - alinhado com XLSX)
    sub_rows = []
    if subscription_summary and subscription_summary.items:
        for item in subscription_summary.items:
            sub_rows.append([
                item.subscription_id,
                item.subscription_name or "—",
                round(item.total_cost_usd, 2),
                round(item.percentage_of_total, 2),
                item.row_count,
                item.max_date.strftime("%d/%m/%Y") if item.max_date else "",
            ])
    sub_csv = _csv_bytes_br(
        ["Subscription ID", "Nome", "Custo Total (R$)", "% do Total", "Registros", "Última Data"],
        sub_rows,
    )

    # 9. Recomendações (alinhado com XLSX)
    # rec_rows already built above from opportunities

    # 10. Recursos (alinhado com XLSX)
    # res_rows already built above from top_services

    res_csv = _csv_bytes_br(
        ["Serviço", "Descrição", "Gasto Mensal (R$)", "% do Total"],
        res_rows,
    )

    rec_csv = _csv_bytes_br(
        ["Prioridade", "Título", "Categoria", "Economia Anual (R$)", "Risco", "Justificativa"],
        rec_rows,
    )

    # Build ZIP
    zip_buf = io.BytesIO()
    with ZipFile(zip_buf, "w", compression=ZIP_DEFLATED) as zf:
        zf.writestr("executive_summary.csv", summary_csv)
        zf.writestr("subscriptions.csv", sub_csv)
        zf.writestr("recommendations.csv", rec_csv)
        zf.writestr("resources.csv", res_csv)
        zf.writestr("governance.csv", gov_combined)
        zf.writestr("sustainability.csv", green_csv)
    return zip_buf.getvalue()


# ---------------------------------------------------------------------------
# Styling constants (professional Brazilian design)
# ---------------------------------------------------------------------------

_BRAND_DARK = "1B2A4A"
_BRAND_ACCENT = "2E86AB"
_BRAND_GOLD = "D4A23B"
_BRAND_GREEN = "4CAF50"
_ROW_ALT = "F0F6FC"
_HEADER_FILL = PatternFill(start_color=_BRAND_DARK, end_color=_BRAND_DARK, fill_type="solid")
_ALT_FILL = PatternFill(start_color=_ROW_ALT, end_color=_ROW_ALT, fill_type="solid")
_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_TITLE_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=16)
_SUBTITLE_FONT = Font(name="Calibri", bold=True, color=_BRAND_DARK, size=12)
_KPI_VALUE_FONT = Font(name="Calibri", bold=True, size=14)
_KPI_LABEL_FONT = Font(name="Calibri", color="555555", size=10)
_THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)
_BRL_FORMAT = 'R$ #,##0.00'
_PCT_FORMAT = '0.0%'
_DATE_FORMAT_BR = 'DD/MM/YYYY'


def _auto_width(ws, min_width: int = 10, max_width: int = 50) -> None:
    for col_cells in ws.columns:
        length = min_width
        for cell in col_cells:
            if cell.value is not None:
                length = max(length, min(len(str(cell.value)) + 4, max_width))
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = length


def _write_table_header(ws, row: int, headers: list[str], num_cols: int) -> None:
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _THIN_BORDER
    # Apply filters
    if num_cols > 0:
        ws.auto_filter.ref = ws.dimensions


def _apply_table_style(ws, start_row: int, end_row: int, num_cols: int, currency_cols: list[int] | None = None, pct_cols: list[int] | None = None) -> None:
    currency_cols = currency_cols or []
    pct_cols = pct_cols or []
    for r in range(start_row, end_row + 1):
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(vertical="center", horizontal="left")
            if c in currency_cols:
                cell.number_format = _BRL_FORMAT
                cell.alignment = Alignment(vertical="center", horizontal="right")
            elif c in pct_cols:
                cell.number_format = _PCT_FORMAT
                cell.alignment = Alignment(vertical="center", horizontal="center")
            if (r - start_row) % 2 == 1:
                if not cell.fill or not cell.fill.start_color or cell.fill.start_color.rgb in (None, "00000000"):
                    cell.fill = _ALT_FILL


def _build_professional_xlsx(
    *,
    dashboard,
    top_services,
    top_teams,
    trend,
    opportunities,
    total_monthly_savings: float,
    total_annual_savings: float,
    subscription_summary: SubscriptionCostSummary,
    gov_summary,
    gov_unowned,
    gov_compliance,
    green_summary,
    green_monthly,
    generated_at: datetime,
    window_days: int,
    filters_json: str,
    org_name: str,
) -> bytes:
    wb = Workbook()

    # Defensive: ensure all optional iterables are never None
    top_services = top_services or []
    top_teams = top_teams or []
    trend = trend or []
    opportunities = opportunities or []
    gov_unowned = gov_unowned or []
    gov_compliance = gov_compliance or []
    green_monthly = green_monthly or []

    # --- Sheet 1: Resumo Executivo ---
    ws = wb.active
    ws.title = "Resumo Executivo"
    _build_executive_summary(
        ws, dashboard, top_services, top_teams, opportunities,
        total_monthly_savings, total_annual_savings,
        generated_at, window_days, org_name,
    )

    # --- Sheet 2: Custos por Subscription ---
    ws2 = wb.create_sheet("Custos por Subscription")
    _build_subscription_sheet(ws2, subscription_summary or None)

    # --- Sheet 3: Custos por Serviço ---
    ws3 = wb.create_sheet("Custos por Serviço")
    _build_services_sheet(ws3, top_services or [], top_teams or [])

    # --- Sheet 4: Tendência de Gastos ---
    ws4 = wb.create_sheet("Tendência de Gastos")
    _build_trend_sheet(ws4, trend or [])

    # --- Sheet 5: Oportunidades de Economia ---
    ws5 = wb.create_sheet("Oportunidades de Economia")
    _build_opportunities_sheet(ws5, opportunities or [], total_monthly_savings, total_annual_savings)

    # --- Sheet 6: Recomendações ---
    ws6 = wb.create_sheet("Recomendações")
    _build_recommendations_sheet(ws6, opportunities or [])

    # --- Sheet 7: Recursos ---
    ws7 = wb.create_sheet("Recursos")
    _build_resources_sheet(ws7, top_services or [])

    # --- Sheet 8: Governança ---
    ws8 = wb.create_sheet("Governança")
    _build_governance_sheet(ws8, gov_summary, gov_unowned or [], gov_compliance or [])

    # --- Sheet 9: Sustentabilidade ---
    ws9 = wb.create_sheet("Sustentabilidade")
    _build_sustainability_sheet(ws9, green_summary, green_monthly or [])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_executive_summary(ws, dashboard, top_services, top_teams, opportunities,
                             total_monthly_savings, total_annual_savings,
                             generated_at, window_days, org_name) -> None:
    # Title row with brand color
    ws.merge_cells("A1:F1")
    ws.row_dimensions[1].height = 40
    title_cell = ws.cell(row=1, column=1, value="CauSium — Relatório FinOps Executivo")
    title_cell.font = _TITLE_FONT
    title_cell.fill = PatternFill(start_color=_BRAND_DARK, end_color=_BRAND_DARK, fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # KPI boxes row
    ws.merge_cells("A3:B3")
    ws.cell(row=3, column=1, value="GASTO MENSAL").font = Font(name="Calibri", bold=True, color=_BRAND_DARK, size=10)
    ws.merge_cells("C3:D3")
    ws.cell(row=3, column=3, value="ECONOMIA MENSAL").font = Font(name="Calibri", bold=True, color=_BRAND_DARK, size=10)
    ws.merge_cells("E3:F3")
    ws.cell(row=3, column=5, value="VARIAÇÃO").font = Font(name="Calibri", bold=True, color=_BRAND_DARK, size=10)

    # KPI values
    ws.merge_cells("A4:B4")
    c = ws.cell(row=4, column=1, value=dashboard.current_month_cost)
    c.font = Font(name="Calibri", bold=True, size=20, color=_BRAND_ACCENT)
    c.number_format = _BRL_FORMAT

    ws.merge_cells("C4:D4")
    c = ws.cell(row=4, column=3, value=total_monthly_savings)
    c.font = Font(name="Calibri", bold=True, size=20, color=_BRAND_GREEN)
    c.number_format = _BRL_FORMAT

    ws.merge_cells("E4:F4")
    c = ws.cell(row=4, column=5, value=dashboard.mom_change_pct / 100.0)
    c.font = Font(name="Calibri", bold=True, size=20, color=_BRAND_ACCENT if dashboard.mom_change_pct <= 0 else "E53935")
    c.number_format = _PCT_FORMAT

    # Report metadata
    row = 6
    ws.cell(row=row, column=1, value="INFORMAÇÕES DO RELATÓRIO").font = _SUBTITLE_FONT
    ws.merge_cells(f"A{row}:F{row}")
    row += 1

    metadata = [
        ("Organização:", org_name),
        ("Período:", f"{window_days} dias"),
        ("Data de Geração:", _format_datetime_br(generated_at) + " UTC"),
    ]
    for label, value in metadata:
        ws.cell(row=row, column=1, value=label).font = Font(name="Calibri", bold=True, color=_BRAND_DARK)
        ws.cell(row=row, column=2, value=value)
        row += 1

    # Indicators section
    row += 1
    ws.cell(row=row, column=1, value="INDICADORES PRINCIPAIS").font = _SUBTITLE_FONT
    ws.merge_cells(f"A{row}:F{row}")
    row += 1

    kpis = [
        ("Gasto Mensal Atual", dashboard.current_month_cost),
        ("Gasto Mês Anterior", dashboard.previous_month_cost),
        ("Variação MoM", dashboard.mom_change_pct / 100.0),
        ("Economia Potencial (Mensal)", total_monthly_savings),
        ("Economia Potencial (Anual)", total_annual_savings),
        ("Contas Cloud Ativas", dashboard.active_accounts),
        ("Oportunidades em Aberto", len(opportunities)),
        ("Eventos (Últimos 7 dias)", dashboard.event_count_7d),
    ]
    for label, value in kpis:
        ws.cell(row=row, column=1, value=label).font = _KPI_LABEL_FONT
        val_cell = ws.cell(row=row, column=2, value=value)
        val_cell.font = _KPI_VALUE_FONT
        if "R$" in label or "Gasto" in label or "Economia" in label:
            val_cell.number_format = _BRL_FORMAT
        elif "%" in label or "Variação" in label:
            val_cell.number_format = _PCT_FORMAT
        row += 1

    # Top 5 Services
    row += 1
    ws.cell(row=row, column=1, value="TOP 5 SERVIÇOS POR GASTO").font = _SUBTITLE_FONT
    ws.merge_cells(f"A{row}:F{row}")
    row += 1
    headers = ["Serviço", "Gasto Mensal (R$)", "% do Total"]
    _write_table_header(ws, row, headers, 3)
    header_row = row
    row += 1
    for svc in top_services[:5]:
        ws.cell(row=row, column=1, value=svc.service)
        ws.cell(row=row, column=2, value=round(svc.cost_usd, 2))
        ws.cell(row=row, column=3, value=round(svc.percentage / 100.0, 3))
        row += 1
    _apply_table_style(ws, header_row + 1, row - 1, 3, currency_cols=[2], pct_cols=[3])

    # Top 5 Teams
    row += 1
    ws.cell(row=row, column=1, value="TOP 5 EQUIPES POR GASTO").font = _SUBTITLE_FONT
    ws.merge_cells(f"A{row}:F{row}")
    row += 1
    _write_table_header(ws, row, ["Equipe", "Gasto Mensal (R$)", "% do Total"], 3)
    header_row = row
    row += 1
    for team in top_teams[:5]:
        ws.cell(row=row, column=1, value=team.service)
        ws.cell(row=row, column=2, value=round(team.cost_usd, 2))
        ws.cell(row=row, column=3, value=round(team.percentage / 100.0, 3))
        row += 1
    _apply_table_style(ws, header_row + 1, row - 1, 3, currency_cols=[2], pct_cols=[3])

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A{header_row}:C{row - 1}"
    _auto_width(ws)


def _build_subscription_sheet(ws, subscription_summary) -> None:
    ws.merge_cells("A1:E1")
    ws.row_dimensions[1].height = 30
    title = ws.cell(row=1, column=1, value="Custos por Subscription")
    title.font = _TITLE_FONT
    title.fill = PatternFill(start_color=_BRAND_DARK, end_color=_BRAND_DARK, fill_type="solid")
    title.alignment = Alignment(horizontal="center", vertical="center")

    # Handle None gracefully
    if subscription_summary is None:
        ws.cell(row=3, column=1, value="Nenhum dado de subscription disponível.")
        ws.cell(row=3, column=1).font = Font(italic=True, color="666666")
        ws.freeze_panes = "A5"
        _auto_width(ws)
        return

    # Summary KPIs
    row = 3
    ws.cell(row=row, column=1, value="Total de Subscriptions:").font = _KPI_LABEL_FONT
    ws.cell(row=row, column=2, value=subscription_summary.subscription_count)
    row += 1
    ws.cell(row=row, column=1, value="Gasto Total:").font = _KPI_LABEL_FONT
    c = ws.cell(row=row, column=2, value=subscription_summary.total_cost_usd)
    c.font = _KPI_VALUE_FONT
    c.number_format = _BRL_FORMAT

    # Subscriptions table
    row += 2
    ws.cell(row=row, column=1, value="DETALHAMENTO POR SUBSCRIPTION").font = _SUBTITLE_FONT
    ws.merge_cells(f"A{row}:E{row}")
    row += 1

    headers = ["Subscription ID", "Nome", "Gasto Mensal (R$)", "Última Data", "% do Total"]
    _write_table_header(ws, row, headers, 5)
    header_row = row
    row += 1

    for item in (subscription_summary.items or []):
        ws.cell(row=row, column=1, value=item.subscription_id)
        ws.cell(row=row, column=2, value=item.subscription_name or "—")
        c = ws.cell(row=row, column=3, value=item.total_cost_usd)
        c.number_format = _BRL_FORMAT
        ws.cell(row=row, column=4, value=_format_date_br(item.max_date) if item.max_date else "Sem dados")
        c = ws.cell(row=row, column=5, value=item.percentage_of_total / 100.0)
        c.number_format = _PCT_FORMAT
        row += 1

    if subscription_summary.items:
        _apply_table_style(ws, header_row + 1, row - 1, 5, currency_cols=[3], pct_cols=[5])
        # Total row
        ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=row, column=1).fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        c = ws.cell(row=row, column=3, value=subscription_summary.total_cost_usd)
        c.font = Font(bold=True)
        c.number_format = _BRL_FORMAT
        c.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

    ws.freeze_panes = "A5"
    _auto_width(ws)


def _build_services_sheet(ws, top_services, top_teams) -> None:
    ws.merge_cells("A1:C1")
    ws.row_dimensions[1].height = 30
    title = ws.cell(row=1, column=1, value="Custos por Serviço")
    title.font = _TITLE_FONT
    title.fill = PatternFill(start_color=_BRAND_DARK, end_color=_BRAND_DARK, fill_type="solid")
    title.alignment = Alignment(horizontal="center", vertical="center")

    # Services table
    row = 3
    ws.cell(row=row, column=1, value="GASTOS POR SERVIÇO").font = _SUBTITLE_FONT
    row += 1
    headers = ["Serviço", "Gasto Mensal (R$)", "% do Total"]
    _write_table_header(ws, row, headers, 3)
    header_row = row
    row += 1
    for svc in top_services:
        ws.cell(row=row, column=1, value=svc.service)
        ws.cell(row=row, column=2, value=round(svc.cost_usd, 2))
        ws.cell(row=row, column=3, value=round(svc.percentage / 100.0, 3))
        row += 1
    svc_end = row - 1
    if top_services:
        _apply_table_style(ws, header_row + 1, svc_end, 3, currency_cols=[2], pct_cols=[3])

    # Services bar chart
    if len(top_services) > 1:
        chart = BarChart()
        chart.type = "bar"
        chart.title = "Top Serviços por Gasto"
        chart.style = 10
        chart.width = 20
        chart.height = 12
        data_ref = Reference(ws, min_col=2, min_row=header_row, max_row=svc_end)
        cats_ref = Reference(ws, min_col=1, min_row=header_row + 1, max_row=svc_end)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.legend = None
        ws.add_chart(chart, "E3")

    # Teams table
    row = svc_end + 3
    ws.cell(row=row, column=1, value="GASTOS POR EQUIPE").font = _SUBTITLE_FONT
    row += 1
    _write_table_header(ws, row, ["Equipe", "Gasto Mensal (R$)", "% do Total"], 3)
    header_row = row
    row += 1
    for team in top_teams:
        ws.cell(row=row, column=1, value=team.service)
        ws.cell(row=row, column=2, value=round(team.cost_usd, 2))
        ws.cell(row=row, column=3, value=round(team.percentage / 100.0, 3))
        row += 1
    if top_teams:
        _apply_table_style(ws, header_row + 1, row - 1, 3, currency_cols=[2], pct_cols=[3])

    ws.freeze_panes = "A3"
    _auto_width(ws)


def _build_trend_sheet(ws, trend) -> None:
    ws.merge_cells("A1:C1")
    ws.row_dimensions[1].height = 30
    title = ws.cell(row=1, column=1, value="Tendência de Gastos Diários")
    title.font = _TITLE_FONT
    title.fill = PatternFill(start_color=_BRAND_DARK, end_color=_BRAND_DARK, fill_type="solid")
    title.alignment = Alignment(horizontal="center", vertical="center")

    row = 3
    headers = ["Data", "Custo Diário (R$)", "Provedor"]
    _write_table_header(ws, row, headers, 3)
    header_row = row
    row += 1
    for t in trend:
        ws.cell(row=row, column=1, value=t.date).number_format = _DATE_FORMAT_BR
        c = ws.cell(row=row, column=2, value=round(t.cost_usd, 2))
        c.number_format = _BRL_FORMAT
        ws.cell(row=row, column=3, value=t.provider or "Todos")
        row += 1
    trend_end = row - 1
    if trend:
        _apply_table_style(ws, header_row + 1, trend_end, 3, currency_cols=[2])

    # Trend line chart
    if len(trend) > 1:
        chart = LineChart()
        chart.title = "Evolução Diária de Gastos"
        chart.style = 10
        chart.y_axis.title = "R$"
        chart.x_axis.title = "Data"
        chart.width = 24
        chart.height = 12
        data_ref = Reference(ws, min_col=2, min_row=header_row, max_row=trend_end)
        cats_ref = Reference(ws, min_col=1, min_row=header_row + 1, max_row=trend_end)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.legend = None
        ws.add_chart(chart, "E3")

    ws.freeze_panes = "A3"
    _auto_width(ws)


def _build_opportunities_sheet(ws, opportunities, total_monthly_savings, total_annual_savings) -> None:
    ws.merge_cells("A1:J1")
    ws.row_dimensions[1].height = 30
    title = ws.cell(row=1, column=1, value="Oportunidades de Economia")
    title.font = _TITLE_FONT
    title.fill = PatternFill(start_color=_BRAND_DARK, end_color=_BRAND_DARK, fill_type="solid")
    title.alignment = Alignment(horizontal="center", vertical="center")

    # Summary row
    ws.merge_cells("A3:B3")
    ws.cell(row=3, column=1, value="Economia Mensal Total:").font = _KPI_LABEL_FONT
    c = ws.cell(row=3, column=3, value=total_monthly_savings)
    c.font = Font(name="Calibri", bold=True, size=16, color=_BRAND_GREEN)
    c.number_format = _BRL_FORMAT

    ws.merge_cells("D3:E3")
    ws.cell(row=3, column=4, value="Economia Anual Total:").font = _KPI_LABEL_FONT
    c = ws.cell(row=3, column=6, value=total_annual_savings)
    c.font = Font(name="Calibri", bold=True, size=16, color=_BRAND_GREEN)
    c.number_format = _BRL_FORMAT

    # Table
    row = 5
    headers = [
        "Oportunidade", "Categoria", "Economia Mensal (R$)", "Economia Anual (R$)",
        "Risco", "Esforço", "Serviço Cloud", "Recurso", "Região", "Status",
    ]
    _write_table_header(ws, row, headers, 10)
    header_row = row
    row += 1

    _RISK_FILLS = {
        "low": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "medium": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "high": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    }

    for op in opportunities:
        ws.cell(row=row, column=1, value=op.title)
        ws.cell(row=row, column=2, value=_format_category(op.category.value))
        c = ws.cell(row=row, column=3, value=round(op.estimated_monthly_savings_usd, 2))
        c.number_format = _BRL_FORMAT
        c = ws.cell(row=row, column=4, value=round(op.estimated_annual_savings_usd, 2))
        c.number_format = _BRL_FORMAT
        risk_cell = ws.cell(row=row, column=5, value=op.risk_level.value.capitalize())
        risk_fill = _RISK_FILLS.get(op.risk_level.value)
        if risk_fill:
            risk_cell.fill = risk_fill
        ws.cell(row=row, column=6, value=op.effort_level.value.capitalize())
        ws.cell(row=row, column=7, value=op.service or "—")
        ws.cell(row=row, column=8, value=op.resource_name or "—")
        ws.cell(row=row, column=9, value=op.region or "—")
        ws.cell(row=row, column=10, value=_format_status(op.status.value))
        row += 1

    if opportunities:
        _apply_table_style(ws, header_row + 1, row - 1, 10, currency_cols=[3, 4])

    # Category pie chart
    category_counts = Counter(_format_category(op.category.value) for op in opportunities)
    if category_counts:
        pie_start = row + 2
        ws.cell(row=pie_start, column=1, value="Categoria").font = _HEADER_FONT
        ws.cell(row=pie_start, column=1).fill = _HEADER_FILL
        ws.cell(row=pie_start, column=2, value="Quantidade").font = _HEADER_FONT
        ws.cell(row=pie_start, column=2).fill = _HEADER_FILL
        r = pie_start + 1
        for cat, cnt in category_counts.most_common():
            ws.cell(row=r, column=1, value=cat)
            ws.cell(row=r, column=2, value=cnt)
            r += 1

        pie = PieChart()
        pie.title = "Oportunidades por Categoria"
        pie.width = 14
        pie.height = 10
        data_ref = Reference(ws, min_col=2, min_row=pie_start, max_row=r - 1)
        cats_ref = Reference(ws, min_col=1, min_row=pie_start + 1, max_row=r - 1)
        pie.add_data(data_ref, titles_from_data=True)
        pie.set_categories(cats_ref)
        ws.add_chart(pie, f"K{pie_start}")

    ws.freeze_panes = "A5"
    _auto_width(ws)


def _build_recommendations_sheet(ws, opportunities) -> None:
    ws.merge_cells("A1:F1")
    ws.row_dimensions[1].height = 30
    title = ws.cell(row=1, column=1, value="Recomendações Prioritárias")
    title.font = _TITLE_FONT
    title.fill = PatternFill(start_color=_BRAND_DARK, end_color=_BRAND_DARK, fill_type="solid")
    title.alignment = Alignment(horizontal="center", vertical="center")

    row = 3
    ws.cell(row=row, column=1, value="RECOMENDAÇÕES BASEADAS EM OPORTUNIDADES DE ECONOMIA").font = _SUBTITLE_FONT
    ws.merge_cells(f"A{row}:F{row}")
    row += 2

    # Priority headers
    headers = ["Prioridade", "Título", "Categoria", "Economia (R$)", "Risco", "Justificativa"]
    _write_table_header(ws, row, headers, 6)
    header_row = row
    row += 1

    priority = 1
    for op in sorted(opportunities, key=lambda x: x.estimated_annual_savings_usd, reverse=True)[:20]:
        ws.cell(row=row, column=1, value=f"#{priority}")
        ws.cell(row=row, column=2, value=op.title)
        ws.cell(row=row, column=3, value=_format_category(op.category.value))
        c = ws.cell(row=row, column=4, value=round(op.estimated_annual_savings_usd, 2))
        c.number_format = _BRL_FORMAT
        ws.cell(row=row, column=5, value=op.risk_level.value.capitalize())

        # Generate justification based on opportunity
        just = f"Implementar {_format_category(op.category.value).lower()}"
        if op.service:
            just += f" no serviço {op.service}"
        if op.region:
            just += f" na região {op.region}"
        just += f". Economia potencial de R$ {op.estimated_annual_savings_usd:,.2f}/ano."
        ws.cell(row=row, column=6, value=just)
        row += 1
        priority += 1

    if opportunities:
        _apply_table_style(ws, header_row + 1, row - 1, 6, currency_cols=[4])

    ws.freeze_panes = "A5"
    _auto_width(ws)


def _build_resources_sheet(ws, top_services) -> None:
    ws.merge_cells("A1:D1")
    ws.row_dimensions[1].height = 30
    title = ws.cell(row=1, column=1, value="Recursos com Maiores Custos")
    title.font = _TITLE_FONT
    title.fill = PatternFill(start_color=_BRAND_DARK, end_color=_BRAND_DARK, fill_type="solid")
    title.alignment = Alignment(horizontal="center", vertical="center")

    row = 3
    ws.cell(row=row, column=1, value="PRINCIPAIS RECURSOS POR SERVIÇO").font = _SUBTITLE_FONT
    ws.merge_cells(f"A{row}:D{row}")
    row += 1

    headers = ["Serviço", "Descrição", "Gasto Mensal (R$)", "% do Total"]
    _write_table_header(ws, row, headers, 4)
    header_row = row
    row += 1

    total_cost = sum(s.cost_usd for s in top_services)
    for svc in top_services:
        pct = (svc.cost_usd / total_cost * 100) if total_cost > 0 else 0
        ws.cell(row=row, column=1, value=svc.service)
        ws.cell(row=row, column=2, value="Recurso principal deste serviço")
        c = ws.cell(row=row, column=3, value=round(svc.cost_usd, 2))
        c.number_format = _BRL_FORMAT
        c = ws.cell(row=row, column=4, value=round(pct / 100.0, 3))
        c.number_format = _PCT_FORMAT
        row += 1

    if top_services:
        _apply_table_style(ws, header_row + 1, row - 1, 4, currency_cols=[3], pct_cols=[4])

    ws.freeze_panes = "A5"
    _auto_width(ws)


def _build_governance_sheet(ws, gov_summary, gov_unowned, gov_compliance) -> None:
    ws.merge_cells("A1:F1")
    ws.row_dimensions[1].height = 30
    title = ws.cell(row=1, column=1, value="Governança Cloud")
    title.font = _TITLE_FONT
    title.fill = PatternFill(start_color=_BRAND_DARK, end_color=_BRAND_DARK, fill_type="solid")
    title.alignment = Alignment(horizontal="center", vertical="center")

    # Summary KPIs
    row = 3
    kpis = [
        ("Total de Recursos", gov_summary.total_resources),
        ("Recursos sem Dono", gov_summary.unowned_resources),
        ("% sem Dono", gov_summary.unowned_pct / 100.0),
        ("Custo sem Dono (R$)", gov_summary.unowned_cost_usd),
        ("Equipes Avaliadas", gov_summary.teams_evaluated),
        ("Compliance Médio (%)", gov_summary.avg_compliance_pct / 100.0),
    ]
    for label, value in kpis:
        ws.cell(row=row, column=1, value=label).font = _KPI_LABEL_FONT
        c = ws.cell(row=row, column=2, value=value)
        c.font = _KPI_VALUE_FONT
        if "R$" in label:
            c.number_format = _BRL_FORMAT
        elif "%" in label:
            c.number_format = _PCT_FORMAT
        row += 1

    # Unowned costs table
    row += 1
    ws.cell(row=row, column=1, value="RECURSOS SEM PROPRIETÁRIO").font = _SUBTITLE_FONT
    ws.merge_cells(f"A{row}:F{row}")
    row += 1
    if gov_unowned:
        headers = ["Serviço", "Recurso", "Região", "Ambiente", "Custo (R$)", "Dias Ativo"]
        _write_table_header(ws, row, headers, 6)
        header_row = row
        row += 1
        for item in gov_unowned:
            ws.cell(row=row, column=1, value=item.service)
            ws.cell(row=row, column=2, value=item.resource_id)
            ws.cell(row=row, column=3, value=item.region)
            ws.cell(row=row, column=4, value=item.environment)
            c = ws.cell(row=row, column=5, value=item.cost_usd)
            c.number_format = _BRL_FORMAT
            ws.cell(row=row, column=6, value=item.days_active)
            row += 1
        _apply_table_style(ws, header_row + 1, row - 1, 6, currency_cols=[5])
    else:
        ws.cell(row=row, column=1, value="Nenhum recurso sem proprietário encontrado.")
        row += 1

    # Label compliance table
    row += 1
    ws.cell(row=row, column=1, value="COMPLIANCE DE TAGS POR EQUIPE").font = _SUBTITLE_FONT
    ws.merge_cells(f"A{row}:D{row}")
    row += 1
    if gov_compliance:
        headers = ["Equipe", "Custo Total (R$)", "Custo sem Tag (R$)", "Compliance (%)"]
        _write_table_header(ws, row, headers, 4)
        header_row = row
        row += 1
        for item in gov_compliance:
            ws.cell(row=row, column=1, value=item.team)
            c = ws.cell(row=row, column=2, value=item.total_cost_usd)
            c.number_format = _BRL_FORMAT
            c = ws.cell(row=row, column=3, value=item.untagged_cost_usd)
            c.number_format = _BRL_FORMAT
            c = ws.cell(row=row, column=4, value=item.compliance_pct / 100.0)
            c.number_format = _PCT_FORMAT
            row += 1
        _apply_table_style(ws, header_row + 1, row - 1, 4, currency_cols=[2, 3], pct_cols=[4])
    else:
        ws.cell(row=row, column=1, value="Dados de compliance não disponíveis.")

    ws.freeze_panes = "A3"
    _auto_width(ws)


def _build_sustainability_sheet(ws, green_summary, green_monthly) -> None:
    ws.merge_cells("A1:D1")
    ws.row_dimensions[1].height = 30
    title = ws.cell(row=1, column=1, value="Sustentabilidade — Emissões de Carbono")
    title.font = _TITLE_FONT
    title.fill = PatternFill(start_color=_BRAND_DARK, end_color=_BRAND_DARK, fill_type="solid")
    title.alignment = Alignment(horizontal="center", vertical="center")

    # Summary KPIs
    row = 3
    kpis = [
        ("Emissões Totais (kgCO2e)", green_summary.total_kg_co2e, "#,##0.0"),
        ("Custo Associado (R$)", green_summary.total_cost_usd, _BRL_FORMAT),
        ("Intensidade Média (gCO2e/R$)", green_summary.intensity_avg, "#,##0.0"),
        ("Variação MoM", (green_summary.mom_delta_pct or 0) / 100.0, _PCT_FORMAT),
        ("Meses Disponíveis", green_summary.months_available, None),
        ("Fonte dos Dados", green_summary.data_source, None),
    ]
    for label, value, fmt in kpis:
        ws.cell(row=row, column=1, value=label).font = _KPI_LABEL_FONT
        c = ws.cell(row=row, column=2, value=value)
        c.font = _KPI_VALUE_FONT
        if fmt:
            c.number_format = fmt
        row += 1

    # Note
    row += 1
    ws.cell(row=row, column=1, value=green_summary.note).font = Font(
        name="Calibri", italic=True, color="666666", size=9
    )

    # Monthly emissions table
    row += 2
    ws.cell(row=row, column=1, value="EMISSÕES MENSAIS").font = _SUBTITLE_FONT
    ws.merge_cells(f"A{row}:D{row}")
    row += 1
    if green_monthly:
        headers = ["Mês", "Emissões (kgCO2e)", "Custo (R$)", "Variação (%)"]
        _write_table_header(ws, row, headers, 4)
        header_row = row
        row += 1
        for m in green_monthly:
            ws.cell(row=row, column=1, value=m.month)
            c = ws.cell(row=row, column=2, value=m.kg_co2e)
            c.number_format = "#,##0.0"
            c = ws.cell(row=row, column=3, value=m.cost_usd)
            c.number_format = _BRL_FORMAT
            c = ws.cell(row=row, column=4, value=(m.delta_pct or 0) / 100.0 if m.delta_pct else None)
            c.number_format = _PCT_FORMAT if m.delta_pct else None
            row += 1
        _apply_table_style(ws, header_row + 1, row - 1, 4, currency_cols=[3], pct_cols=[4])

        # Emissions line chart
        if len(green_monthly) > 1:
            chart = LineChart()
            chart.title = "Evolução de Emissões Mensais"
            chart.style = 10
            chart.y_axis.title = "kgCO2e"
            chart.width = 18
            chart.height = 10
            data_ref = Reference(ws, min_col=2, min_row=header_row, max_row=row - 1)
            cats_ref = Reference(ws, min_col=1, min_row=header_row + 1, max_row=row - 1)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            chart.legend = None
            ws.add_chart(chart, f"F{header_row}")
    else:
        ws.cell(row=row, column=1, value="Dados de emissões não disponíveis para este período.")

    ws.freeze_panes = "A3"
    _auto_width(ws)