from __future__ import annotations

from zipfile import ZipFile
import io

from openpyxl import load_workbook

from app.domains.economics.export_runtime import build_xlsx_workbook, _build_sustainability_sheet
from app.domains.green.schemas import EmissionsMonthRowOut, GreenSummaryOut


def test_build_xlsx_workbook_contains_expected_parts():
    workbook = build_xlsx_workbook(
        [
            ("Summary", [["key", "value"], ["current_month_cost", 123.45]]),
            ("Top Services", [["service", "cost_usd"], ["Compute", 42.0]]),
        ]
    )

    with ZipFile(io.BytesIO(workbook)) as archive:
        names = set(archive.namelist())
        assert "[Content_Types].xml" in names
        assert "xl/workbook.xml" in names
        assert "xl/worksheets/sheet1.xml" in names
        assert "xl/worksheets/sheet2.xml" in names

        sheet1 = archive.read("xl/worksheets/sheet1.xml").decode("utf-8")
        assert "current_month_cost" in sheet1
        assert "123.45" in sheet1


def test_sustainability_sheet_with_delta_pct_none():
    """Test that sustainability sheet handles delta_pct=None without breaking.

    Regression test for: TypeError: openpyxl.styles.numbers.NumberFormat.formatCode
    should be <class 'str'> but value is <class 'NoneType'>
    """
    from openpyxl import Workbook

    # Create mock data with delta_pct=None (first month has no previous to compare)
    green_monthly = [
        EmissionsMonthRowOut(month="2026-01", kg_co2e=100.0, cost_usd=50.0, delta_pct=None),
        EmissionsMonthRowOut(month="2026-02", kg_co2e=110.0, cost_usd=55.0, delta_pct=10.0),
        EmissionsMonthRowOut(month="2026-03", kg_co2e=105.0, cost_usd=52.5, delta_pct=-4.5),
    ]

    green_summary = GreenSummaryOut(
        total_kg_co2e=315.0,
        total_cost_usd=157.5,
        intensity_avg=2.0,
        mom_delta_pct=5.0,
        months_available=3,
        data_source="mock",
        note="Test data",
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Sustentabilidade"

    # This should NOT raise TypeError
    _build_sustainability_sheet(ws, green_summary, green_monthly)

    # Verify workbook can be saved and loaded (validates openpyxl compatibility)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Load the workbook to verify it's valid
    loaded_wb = load_workbook(buffer)
    loaded_ws = loaded_wb.active

    # Verify data is present
    assert loaded_ws["A1"].value == "Sustentabilidade — Emissões de Carbono"

    # Find the emissions table and verify structure
    # The monthly data should be in the table
    found_months = []
    for row in loaded_ws.iter_rows(min_row=1, max_row=50, values_only=True):
        if row[0] and isinstance(row[0], str) and row[0].startswith("20"):
            found_months.append(row[0])

    assert "2026-01" in found_months
    assert "2026-02" in found_months
    assert "2026-03" in found_months